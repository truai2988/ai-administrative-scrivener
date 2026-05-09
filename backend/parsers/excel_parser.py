"""
Excel パーサー

openpyxl を用いてExcelファイル（.xlsx）からテキストと表データを抽出する。
セル結合情報やシート名を保持し、レイアウト情報をなるべく残す。
"""

import io
import logging
from typing import Any

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)


def extract_excel_text(file_bytes: bytes, filename: str) -> str:
    """
    Excelファイルのバイト列を受け取り、構造化テキストに変換する。

    - 各シートのシート名をセクションヘッダーとして出力
    - セル結合情報を保持しつつ、行ごとにテキストを結合
    - 空行が続く場合はスキップして可読性を向上
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
    except Exception as e:
        logger.error(f"openpyxl でのファイル読み込みに失敗: {filename} - {e}")
        # フォールバック: pandas で読み込む
        return _extract_with_pandas(file_bytes, filename)

    sections: list[str] = []
    sections.append(f"=== ファイル: {filename} ===\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sections.append(f"\n--- シート: {sheet_name} ---")

        # 結合セルの情報を取得
        merged_cells_map: dict[tuple[int, int], str] = {}
        for merged_range in ws.merged_cells.ranges:
            top_left_value = ws.cell(merged_range.min_row, merged_range.min_col).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if (row, col) != (merged_range.min_row, merged_range.min_col):
                        merged_cells_map[(row, col)] = str(top_left_value) if top_left_value else ""

        # 最大行・列を制限（OOMクラッシュ防止のため）
        actual_max_row = min(ws.max_row, 2000) if ws.max_row else 2000
        actual_max_col = min(ws.max_column, 50) if ws.max_column else 50

        consecutive_empty = 0
        for row in ws.iter_rows(min_row=1, max_row=actual_max_row, max_col=actual_max_col):
            cells: list[str] = []
            for cell in row:
                coord = (cell.row, cell.column)
                if coord in merged_cells_map:
                    # 結合セルの一部 → 表示上はスキップ（重複を避ける）
                    continue
                value = cell.value
                if value is not None:
                    cells.append(str(value).strip())

            line = " | ".join(cells)
            if line.strip():
                consecutive_empty = 0
                sections.append(f"  {line}")
            else:
                consecutive_empty += 1
                if consecutive_empty <= 2:  # 空行は最大2行まで
                    sections.append("")

    wb.close()
    return "\n".join(sections)


def _extract_with_pandas(file_bytes: bytes, filename: str) -> str:
    """pandas フォールバック: DataFrameとして読み込み、テキスト化する"""
    try:
        xls = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
    except Exception as e:
        logger.error(f"pandas でのファイル読み込みにも失敗: {filename} - {e}")
        return f"=== ファイル: {filename} (読み込みエラー) ===\n"

    sections: list[str] = [f"=== ファイル: {filename} ===\n"]

    for sheet_name in xls.sheet_names:
        # OOMクラッシュ防止のため、最大2000行に制限
        df = pd.read_excel(xls, sheet_name=sheet_name, header=None, dtype=str, nrows=2000)
        df = df.fillna("")
        sections.append(f"\n--- シート: {sheet_name} ---")

        # 列数を最大50列に制限
        df = df.iloc[:, :50]

        for _, row_data in df.iterrows():
            values = [str(v).strip() for v in row_data if str(v).strip()]
            if values:
                sections.append(f"  {' | '.join(values)}")

    return "\n".join(sections)
